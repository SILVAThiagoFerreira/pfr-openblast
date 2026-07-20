import sys
from pathlib import Path
from tempfile import TemporaryDirectory

import pandas as pd
import unittest
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
SRC = ROOT / "src"
if str(SRC) not in sys.path:
    sys.path.insert(0, str(SRC))

from pfr.config import load_config, normalize_config  # noqa: E402
from pfr.export import export_workbook  # noqa: E402
from pfr.processing import build_output_frame, extract_blast_datetime, extract_plan_id  # noqa: E402


class ProcessingTest(unittest.TestCase):
    def test_plan_id_for_pp230426_b_is_resolved_from_fallback_and_matches_histo_block(self):
        cfg = normalize_config(load_config(ROOT / "config_PP230426_B.yaml"), ROOT)
        histo_content = """[BlastingPlan]2026/06/19-12:10:50;82;+38.7
DBD0233;97
 PU1247;24;26FA5;PP230426_B;;4;2;1;
  1028/104D10D4/0/0
-
[Fire]2026/06/19-12:10:52;81;+38.2
DBD0233;94
-
"""
        with TemporaryDirectory() as tmp:
            path = Path(tmp) / "HISTO-TESTE.txt"
            path.write_text(histo_content, encoding="utf-8")

            plan_id = extract_plan_id(None, (path,), cfg)
            blast_date, blast_time = extract_blast_datetime((path,), plan_id)

        self.assertEqual(plan_id, "230426_B")
        self.assertEqual(blast_date, "19/06/2026")
        self.assertEqual(blast_time, "12:10:52")

    def test_fallback_plan_id_is_not_extracted_from_pdf(self):
        cfg = normalize_config(load_config(ROOT / "config.yaml"), ROOT)
        with TemporaryDirectory() as tmp:
            pdf = Path(tmp) / "PP.pdf"
            pdf.write_bytes(b"not a real pdf")
            self.assertEqual(extract_plan_id(None, (), cfg), "320526")

    def test_missing_plan_fire_event_fails_instead_of_using_current_time(self):
        history = "[BlastingPlan]2026/05/13-11:50:28;84;+37.0\nPP999999\n"
        with TemporaryDirectory() as tmp:
            path = Path(tmp) / "HISTO.txt"
            path.write_text(history, encoding="utf-8")
            with self.assertRaisesRegex(ValueError, "(?i)não foi encontrado.*plano"):
                extract_blast_datetime((path,), "320526")

    def test_blast_datetime_is_tied_to_plan_block(self):
        history = """[BlastingPlan]2026/05/13-06:05:55;88;+26.1
DBD0061;99
 PU2759;1;563A9;PP;;4;3;1;
  5/90009D95/0/10
-
[Fire]2026/05/13-06:08:03;88;+25.8
DBD0061;99
-
[BlastingPlan]2026/05/13-11:50:28;84;+37.0
DBD0061;97
 PU2431;162;4BFBF;PP320526;A;2;2;1;PP320526 _ TEMPORIZACAO _ O_Pitblast
  1028/104D10D4/0/0
-
[Fire]2026/05/13-12:01:43;81;+38.2
DBD0061;94
-
[BlastingPlan]2026/05/19-12:12:41;82;+38.7
DBD0233;97
 PU1247;24;26FA5;DRE-1926;;4;2;1;
-
[Fire]2026/05/19-12:14:13;81;+39.2
DBD0233;97
"""
        with TemporaryDirectory() as tmp:
            path = Path(tmp) / "HISTO-TESTE.txt"
            path.write_text(history, encoding="utf-8")

            blast_date, blast_time = extract_blast_datetime((path,), "320526")

        self.assertEqual(blast_date, "13/05/2026")
        self.assertEqual(blast_time, "12:01:43")

    def test_blast_datetime_allows_detonation_month_to_differ_from_plan_month(self):
        history = """[BlastingPlan]2026/07/16-12:29:07;84;+34.3
 PU48;368;1853;PP290726;A;2;2;1;PP290726_D _ TEMPORIZACAO
-
[Fire]2026/07/16-12:32:49;83;+33.5
"""
        with TemporaryDirectory() as tmp:
            path = Path(tmp) / "HISTO-TESTE.txt"
            path.write_text(history, encoding="utf-8")
            blast_date, blast_time = extract_blast_datetime((path,), "290426")

        self.assertEqual(blast_date, "16/07/2026")
        self.assertEqual(blast_time, "12:32:49")

    def test_blast_datetime_does_not_match_a_different_plan_number(self):
        history = """[BlastingPlan]2026/07/16-12:29:06;84;+34.3
 PU2759;257;563BC;PP400726;A;2;3;1;PP400726
-
[Fire]2026/07/16-12:32:49;83;+33.5
"""
        with TemporaryDirectory() as tmp:
            path = Path(tmp) / "HISTO-TESTE.txt"
            path.write_text(history, encoding="utf-8")
            with self.assertRaisesRegex(ValueError, "(?i)não foi encontrado.*290426"):
                extract_blast_datetime((path,), "290426")

    def test_blast_datetime_rejects_multiple_compatible_blocks(self):
        history = """[BlastingPlan]2026/07/16-12:29:06;84;+34.3
 PU48;368;1853;PP290726;A;2;2;1;PP290726
-
[Fire]2026/07/16-12:32:49;83;+33.5
[BlastingPlan]2026/07/16-13:29:06;84;+34.3
 PU49;368;1854;PP290726;A;2;2;1;PP290726
-
[Fire]2026/07/16-13:32:49;83;+33.5
"""
        with TemporaryDirectory() as tmp:
            path = Path(tmp) / "HISTO-TESTE.txt"
            path.write_text(history, encoding="utf-8")
            with self.assertRaisesRegex(ValueError, "(?i)múltiplos blocos"):
                extract_blast_datetime((path,), "290426")

    def test_missing_detonating_time_is_interpolated(self):
        cfg = normalize_config(load_config(ROOT / "config.yaml"), ROOT)
        merged = pd.DataFrame(
            {
                "Number": [1, 2, 3],
                "X": [10.0, 11.0, 12.0],
                "Y": [20.0, 21.0, 22.0],
                "Z": [30.0, 31.0, 32.0],
                "Z_Toe": [29.0, 30.0, 31.0],
                "p_length": [9.0, 9.0, 9.0],
                "r_length": [9.0, 9.0, 9.0],
                "p_stemming": [3.0, 3.0, 3.0],
                "r_stemming": [3.0, 3.0, 3.0],
                "p_subdrilling": [0.6, 0.6, 0.6],
                "r_subdrilling": [0.6, 0.6, 0.6],
                "r_azimuth": [0.0, 0.0, 0.0],
                "r_angle": [0.0, 0.0, 0.0],
                "p_explosive": [10.0, 10.0, 10.0],
                "r_explosive": [10.0, 10.0, 10.0],
                "Diameter_m": [0.127, 0.127, 0.127],
                "DetonatingTime": [1000.0, float("nan"), 1300.0],
            }
        )

        data = build_output_frame(merged, "1234567", "30/04/2026", "06:04:41", cfg)

        self.assertEqual(data["tempo detonacao (ms)"].tolist(), [1000, 1150, 1300])
        self.assertEqual(str(data["tempo detonacao (ms)"].dtype), "Int64")
        self.assertEqual(merged.attrs["imputed_detonating_time_count"], 1)

    def test_missing_detonating_time_is_unique_and_monotonic(self):
        cfg = normalize_config(load_config(ROOT / "config.yaml"), ROOT)
        merged = pd.DataFrame(
            {
                "Number": [1, 2, 3, 4],
                "X": [10.0, 11.0, 12.0, 13.0],
                "Y": [20.0, 21.0, 22.0, 23.0],
                "Z": [30.0, 31.0, 32.0, 33.0],
                "Z_Toe": [29.0, 30.0, 31.0, 32.0],
                "p_length": [9.0, 9.0, 9.0, 9.0],
                "r_length": [9.0, 9.0, 9.0, 9.0],
                "p_stemming": [3.0, 3.0, 3.0, 3.0],
                "r_stemming": [3.0, 3.0, 3.0, 3.0],
                "p_subdrilling": [0.6, 0.6, 0.6, 0.6],
                "r_subdrilling": [0.6, 0.6, 0.6, 0.6],
                "r_azimuth": [0.0, 0.0, 0.0, 0.0],
                "r_angle": [0.0, 0.0, 0.0, 0.0],
                "p_explosive": [10.0, 10.0, 10.0, 10.0],
                "r_explosive": [10.0, 10.0, 10.0, 10.0],
                "Diameter_m": [0.127, 0.127, 0.127, 0.127],
                "DetonatingTime": [1000.0, float("nan"), float("nan"), 1005.0],
            }
        )

        data = build_output_frame(merged, "1234567", "30/04/2026", "06:04:41", cfg)

        self.assertEqual(data["tempo detonacao (ms)"].tolist(), [1000, 1001, 1003, 1005])
        self.assertEqual(len(set(data["tempo detonacao (ms)"].tolist())), 4)
        self.assertEqual(merged.attrs["imputed_detonating_time_count"], 2)

    def test_stemming_variation_is_deterministic_and_bounded(self):
        cfg = normalize_config(load_config(ROOT / "config.yaml"), ROOT)
        merged = pd.DataFrame(
            {
                "Number": [10, 11, 12],
                "X": [10.0, 11.0, 12.0],
                "Y": [20.0, 21.0, 22.0],
                "Z": [30.0, 31.0, 32.0],
                "Z_Toe": [29.0, 30.0, 31.0],
                "p_length": [9.0, 9.0, 9.0],
                "r_length": [9.0, 9.0, 9.0],
                "p_stemming": [3.0, 3.0, 3.0],
                "r_stemming": [3.0, 3.0, 3.0],
                "p_subdrilling": [0.6, 0.6, 0.6],
                "r_subdrilling": [0.6, 0.6, 0.6],
                "r_azimuth": [0.0, 0.0, 0.0],
                "r_angle": [0.0, 0.0, 0.0],
                "p_explosive": [10.0, 10.0, 10.0],
                "r_explosive": [10.0, 10.0, 10.0],
                "Diameter_m": [0.127, 0.127, 0.127],
                "DetonatingTime": [1000.0, 1100.0, 1200.0],
            }
        )

        first = build_output_frame(merged.copy(), "1234567", "30/04/2026", "06:04:41", cfg)
        second = build_output_frame(merged.copy(), "1234567", "30/04/2026", "06:04:41", cfg)

        self.assertEqual(first["tampao realizado"].tolist(), second["tampao realizado"].tolist())
        for value in first["tampao realizado"]:
            self.assertGreaterEqual(value, 2.88)
            self.assertLessEqual(value, 3.12)

    def test_stemming_realized_is_exported_with_one_decimal(self):
        cfg = normalize_config(load_config(ROOT / "config.yaml"), ROOT)
        merged = pd.DataFrame(
            {
                "Number": [1],
                "X": [10.0],
                "Y": [20.0],
                "Z": [30.0],
                "Z_Toe": [29.0],
                "p_length": [9.0],
                "r_length": [9.0],
                "p_stemming": [3.0],
                "r_stemming": [3.04],
                "p_subdrilling": [0.6],
                "r_subdrilling": [0.6],
                "r_azimuth": [0.0],
                "r_angle": [0.0],
                "p_explosive": [10.0],
                "r_explosive": [10.0],
                "Diameter_m": [0.127],
                "DetonatingTime": [1000.0],
            }
        )

        data = build_output_frame(merged, "1234567", "30/04/2026", "06:04:41", cfg)

        self.assertEqual(data["tampao realizado"].tolist(), [3.0])
        with TemporaryDirectory() as tmp:
            output = Path(tmp) / "saida.xlsx"
            summary = pd.DataFrame([["Plano", "1234567"]], columns=["Campo", "Valor"])
            export_workbook(output, data, summary, cfg)
            wb = load_workbook(output)
            ws = wb[cfg["business"]["data_sheet_name"]]
            self.assertEqual(ws["P2"].number_format, "0.0")
            self.assertEqual(ws["Q2"].number_format, "0.0")

    def test_zero_charges_are_redistributed_without_changing_extremes(self):
        cfg = normalize_config(load_config(ROOT / "config.yaml"), ROOT)
        cfg["business"]["redistribute_zero_charges"] = True
        cfg["business"]["charge_total_target_kg"] = 100.0
        cfg["business"]["charge_zero_fill_min_kg"] = 0.5
        merged = pd.DataFrame(
            {
                "Number": [1, 2, 3, 4, 5],
                "X": [10.0, 11.0, 12.0, 13.0, 14.0],
                "Y": [20.0, 21.0, 22.0, 23.0, 24.0],
                "Z": [30.0, 31.0, 32.0, 33.0, 34.0],
                "Z_Toe": [29.0, 30.0, 31.0, 32.0, 33.0],
                "p_length": [9.0, 9.0, 9.0, 9.0, 9.0],
                "r_length": [9.0, 9.0, 9.0, 9.0, 9.0],
                "p_stemming": [3.0, 3.0, 3.0, 3.0, 3.0],
                "r_stemming": [3.0, 3.0, 3.0, 3.0, 3.0],
                "p_subdrilling": [0.6, 0.6, 0.6, 0.6, 0.6],
                "r_subdrilling": [0.6, 0.6, 0.6, 0.6, 0.6],
                "r_azimuth": [0.0, 0.0, 0.0, 0.0, 0.0],
                "r_angle": [0.0, 0.0, 0.0, 0.0, 0.0],
                "p_explosive": [10.0, 10.0, 10.0, 10.0, 10.0],
                "r_explosive": [5.0, 0.0, 20.0, 0.0, 60.0],
                "Diameter_m": [0.127, 0.127, 0.127, 0.127, 0.127],
                "DetonatingTime": [1000.0, 1100.0, 1200.0, 1300.0, 1400.0],
            }
        )

        data = build_output_frame(merged, "1234567", "30/04/2026", "06:04:41", cfg)
        charges = data["cargas realizadas"].astype(float).tolist()

        self.assertAlmostEqual(sum(charges), 85.0, places=3)
        self.assertGreater(charges[1], 0.0)
        self.assertGreater(charges[3], 0.0)
        self.assertEqual(charges[0], 5.0)
        self.assertEqual(charges[4], 60.0)

    def test_charge_total_target_is_enforced_without_changing_extremes(self):
        cfg = normalize_config(load_config(ROOT / "config.yaml"), ROOT)
        cfg["business"]["enforce_charge_total_target"] = True
        cfg["business"]["charge_total_target_kg"] = 100.0
        merged = pd.DataFrame(
            {
                "Number": [1, 2, 3, 4],
                "X": [10.0, 11.0, 12.0, 13.0],
                "Y": [20.0, 21.0, 22.0, 23.0],
                "Z": [30.0, 31.0, 32.0, 33.0],
                "Z_Toe": [29.0, 30.0, 31.0, 32.0],
                "p_length": [9.0, 9.0, 9.0, 9.0],
                "r_length": [9.0, 9.0, 9.0, 9.0],
                "p_stemming": [3.0, 3.0, 3.0, 3.0],
                "r_stemming": [3.0, 3.0, 3.0, 3.0],
                "p_subdrilling": [0.6, 0.6, 0.6, 0.6],
                "r_subdrilling": [0.6, 0.6, 0.6, 0.6],
                "r_azimuth": [0.0, 0.0, 0.0, 0.0],
                "r_angle": [0.0, 0.0, 0.0, 0.0],
                "p_explosive": [10.0, 10.0, 10.0, 10.0],
                "r_explosive": [5.0, 10.0, 15.0, 60.0],
                "Diameter_m": [0.127, 0.127, 0.127, 0.127],
                "DetonatingTime": [1000.0, 1100.0, 1200.0, 1300.0],
            }
        )

        data = build_output_frame(merged, "1234567", "30/04/2026", "06:04:41", cfg)
        charges = data["cargas realizadas"].astype(float).tolist()

        self.assertAlmostEqual(sum(charges), 100.0, places=3)
        self.assertEqual(charges[0], 5.0)
        self.assertEqual(charges[3], 60.0)
        self.assertGreaterEqual(min(charges), 5.0)
        self.assertLessEqual(max(charges), 60.0)
        self.assertTrue(merged.attrs["charge_total_adjusted"])

    def test_validate_output_blocks_inflated_charge(self):
        from pfr.validation import validate_output

        cfg = normalize_config(load_config(ROOT / "config.yaml"), ROOT)
        input_charges = pd.Series([5.0, 10.0, 178.6, 60.0])

        good = pd.DataFrame({
            "cargas realizadas": [5.0, 10.0, 178.0, 60.0],
            "tempo detonacao (ms)": [1000, 1100, 1200, 1300],
        })
        validate_output(good, input_charges, cfg)

        bad = pd.DataFrame({
            "cargas realizadas": [5.0, 10.0, 3550.0, 60.0],
            "tempo detonacao (ms)": [1000, 1100, 1200, 1300],
        })
        with self.assertRaises(ValueError):
            validate_output(bad, input_charges, cfg)
