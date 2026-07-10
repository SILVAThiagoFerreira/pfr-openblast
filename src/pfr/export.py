from pathlib import Path

import pandas as pd
from openpyxl import load_workbook


def export_workbook(path: Path, data: pd.DataFrame, summary: pd.DataFrame, cfg: dict) -> None:
    data_sheet = cfg["business"]["data_sheet_name"]
    summary_sheet = cfg["business"]["summary_sheet_name"]
    stemming_decimals = int(cfg.get("export", {}).get("decimal_places_stemming_realized", 1))
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        data.to_excel(writer, sheet_name=data_sheet, index=False)
        summary.to_excel(writer, sheet_name=summary_sheet, index=False)

    wb = load_workbook(path)
    ws = wb[data_sheet]
    widths = {"A": 14, "B": 12, "C": 12, "D": 12, "E": 10, "F": 12, "G": 12, "H": 12, "I": 12, "J": 16, "K": 16, "L": 12, "M": 12, "N": 14, "O": 14, "P": 14, "Q": 14, "R": 12, "S": 10, "T": 14}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    stemming_format = f"0.{ '0' * stemming_decimals }" if stemming_decimals > 0 else "0"
    for row in range(2, ws.max_row + 1):
        ws[f"P{row}"].number_format = stemming_format
        ws[f"Q{row}"].number_format = stemming_format
    wb[summary_sheet].column_dimensions["A"].width = 28
    wb[summary_sheet].column_dimensions["B"].width = 42
    wb.save(path)
